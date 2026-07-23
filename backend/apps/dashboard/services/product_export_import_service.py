import logging
import os
from typing import Dict, Any, List, Optional
from datetime import datetime
from decimal import Decimal

import requests
from django.db import transaction
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO

from core.product.services import ProductService
from core.models import Product, ProductCategory, ProductImage, Attachment, ProductField, ProductFormula, FieldDictionary

logger = logging.getLogger('dashboard.services.product_export_import')


class ProductExportImportService:
    """
    سرویس استخراج و ایمپورت محصولات به/از فرمت Excel
    """
    
    def __init__(self):
        self._product_service = ProductService()
    
    # ==========================================
    # EXPORT FUNCTIONALITY
    # ==========================================
    
    def export_products_to_excel(self, product_ids: Optional[List[int]] = None) -> Dict[str, Any]:
        """
        استخراج محصولات به فرمت Excel
        
        Args:
            product_ids: لیست ID محصولات برای استخراج (اگر None یا خالی باشد، همه محصولات استخراج می‌شوند)
        
        Returns:
            Dict containing:
                - file_path: مسیر فایل ذخیره شده
                - file_name: نام فایل
                - product_count: تعداد محصولات استخراج شده
        """
        try:
            # دریافت محصولات
            if product_ids and len(product_ids) > 0:
                products = Product.objects.filter(
                    id__in=product_ids
                ).prefetch_related(
                    'category_relations__category',
                    'fields__field_dict',
                    'fields__choices__choice_dict',
                    'formulas',
                    'product_image',
                    'product_attachment'
                )
            else:
                products = Product.objects.all().prefetch_related(
                    'category_relations__category',
                    'fields__field_dict',
                    'fields__choices__choice_dict',
                    'formulas',
                    'product_image',
                    'product_attachment'
                )
            
            products = products.order_by('-created_at')
            product_count = products.count()
            
            if product_count == 0:
                return {
                    'success': False,
                    'message': 'هیچ محصولی برای استخراج یافت نشد.',
                    'file_path': None,
                    'product_count': 0
                }
            
            # ایجاد Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # استایل‌ها
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ===== Sheet 1: محصولات =====
            headers = [
                "ID", "نام محصول", "کد محصول", "توضیحات", "قیمت", 
                "قیمت نمایشی", "قیمت به ازای", "دارای تیراژ", "فعال",
                "دسته‌بندی‌ها", "تاریخ ایجاد"
            ]
            
            # نوشتن هدرها
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # نوشتن داده‌های محصولات
            for row_num, product in enumerate(products, 2):
                categories = self._get_product_categories_text(product)
                
                row_data = [
                    product.id,
                    product.name,
                    product.code or "",
                    product.description or "",
                    float(product.price) if product.price else 0,
                    float(product.show_price) if product.show_price else 0,
                    product.price_per_unit or 0,
                    "بله" if product.has_quantity else "خیر",
                    "بله" if product.is_active else "خیر",
                    categories,
                    product.created_at.strftime('%Y-%m-%d %H:%M') if product.created_at else ""
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = cell_alignment
                    cell.border = border
            
            # تنظیم عرض ستون‌ها
            column_widths = [8, 30, 20, 50, 15, 15, 15, 12, 10, 40, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 2: فیلدهای محصولات =====
            if products.exists():
                ws_fields = wb.create_sheet("Product Fields")
                field_headers = [
                    "Product ID", "Product Name", "Field ID", "Field Title", 
                    "Field Type", "Value", "Required", "Order"
                ]
                
                for col_num, header in enumerate(field_headers, 1):
                    cell = ws_fields.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                row_num = 2
                for product in products:
                    for field in product.fields.all():
                        field_data = [
                            product.id,
                            product.name,
                            field.id,
                            field.field_dict.title,
                            field.field_dict.field_type,
                            float(field.numeric_value) if field.numeric_value is not None else 0,
                            "بله" if field.is_required else "خیر",
                            field.order
                        ]
                        
                        for col_num, value in enumerate(field_data, 1):
                            cell = ws_fields.cell(row=row_num, column=col_num)
                            cell.value = value
                            cell.alignment = cell_alignment
                            cell.border = border
                        row_num += 1
                
                # تنظیم عرض ستون‌ها
                field_column_widths = [12, 30, 10, 25, 15, 15, 10, 8]
                for i, width in enumerate(field_column_widths, 1):
                    ws_fields.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 3: فرمول‌ها =====
            if products.exists():
                ws_formulas = wb.create_sheet("Formulas")
                formula_headers = [
                    "Product ID", "Product Name", "Formula ID", "Title",
                    "Condition", "Calculation Expression"
                ]
                
                for col_num, header in enumerate(formula_headers, 1):
                    cell = ws_formulas.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                row_num = 2
                for product in products:
                    for formula in product.formulas.all():
                        formula_data = [
                            product.id,
                            product.name,
                            formula.id,
                            formula.title,
                            formula.condition_expression or "",
                            formula.calculation_expression
                        ]
                        
                        for col_num, value in enumerate(formula_data, 1):
                            cell = ws_formulas.cell(row=row_num, column=col_num)
                            cell.value = value
                            cell.alignment = cell_alignment
                            cell.border = border
                        row_num += 1
                
                formula_column_widths = [12, 30, 10, 25, 30, 50]
                for i, width in enumerate(formula_column_widths, 1):
                    ws_formulas.column_dimensions[get_column_letter(i)].width = width
            
            # ذخیره فایل
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"products_export_{timestamp}.xlsx"
            file_path = f"exports/products/{file_name}"
            
            # اطمینان از وجود پوشه
            exports_dir = os.path.join(default_storage.location, 'exports', 'products')
            os.makedirs(exports_dir, exist_ok=True)
            
            # ذخیره در حافظه
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # ذخیره فایل
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
            default_storage.save(file_path, buffer)
            
            logger.info(f"Successfully exported {product_count} products to {file_path}")
            
            return {
                'success': True,
                'message': f'{product_count} محصول با موفقیت استخراج شد.',
                'file_path': file_path,
                'file_name': file_name,
                'product_count': product_count
            }
            
        except Exception as e:
            logger.error(f"Error exporting products: {str(e)}", exc_info=True)
            return {
                'success': False,
                'message': f'خطا در استخراج محصولات: {str(e)}',
                'file_path': None,
                'product_count': 0
            }
    
    def _get_product_categories_text(self, product: Product) -> str:
        """تبدیل دسته‌بندی‌های محصول به متن"""
        categories = []
        for rel in product.category_relations.all().order_by('-is_primary', 'priority'):
            cat = rel.category
            cat_text = cat.name
            if rel.is_primary:
                cat_text += " (اصلی)"
            categories.append(cat_text)
        return " | ".join(categories) if categories else "بدون دسته‌بندی"
    
    # ==========================================
    # IMPORT FUNCTIONALITY (Enhanced with Fields & Formulas)
    # ==========================================
    
    def import_products_from_excel(self, file_path: str, user=None) -> Dict[str, Any]:
        """
        ایمپورت محصولات از فایل Excel (همراه با فیلدها، فرمول‌ها و عکس‌ها)
        
        Args:
            file_path: مسیر فایل Excel
            user: کاربری که درخواست ایمپورت را داده (برای فیلد user در محصول)
        
        Returns:
            Dict containing:
                - success: bool
                - imported_count: تعداد محصولات ایمپورت شده
                - failed_count: تعداد محصولات ناموفق
                - errors: لیست خطاها
        """
        try:
            from openpyxl import load_workbook
            
            # بارگذاری فایل
            if not default_storage.exists(file_path):
                return {
                    'success': False,
                    'message': 'فایل یافت نشد.',
                    'imported_count': 0,
                    'failed_count': 0,
                    'errors': ['فایل آپلود شده وجود ندارد.']
                }
            
            with default_storage.open(file_path, 'rb') as f:
                wb = load_workbook(f, data_only=True)
            
            # خواندن Sheetهای مختلف
            products_ws = wb["Products"] if "Products" in wb.sheetnames else wb.active
            fields_ws = wb["Product Fields"] if "Product Fields" in wb.sheetnames else None
            formulas_ws = wb["Formulas"] if "Formulas" in wb.sheetnames else None
            images_ws = wb["Images"] if "Images" in wb.sheetnames else None
            
            # خواندن داده‌ها
            imported_count = 0
            failed_count = 0
            errors = []
            temp_to_real_ids = {}  # نگاشت ID موقت به ID واقعی
            
            with transaction.atomic():
                # ===== مرحله ۱: ایجاد محصولات =====
                product_headers = [cell.value for cell in products_ws[1]]
                product_id_col = product_headers.index('ID') if 'ID' in product_headers else None
                
                for row_num, row in enumerate(products_ws.iter_rows(min_row=2, values_only=True), 2):
                    if not any(row):
                        continue
                    
                    try:
                        product_data = dict(zip(product_headers, row))
                        
                        # اعتبارسنجی
                        if not product_data.get('نام محصول'):
                            errors.append(f"سطر {row_num}: نام محصول الزامی است.")
                            failed_count += 1
                            continue
                        
                        # ایجاد محصول
                        product = Product(
                            user=user,  # کاربر ایجاد کننده
                            name=product_data.get('نام محصول', '').strip(),
                            description=product_data.get('توضیحات', '') or '',
                            price=Decimal(str(product_data.get('قیمت', 0) or 0)),
                            show_price=Decimal(str(product_data.get('قیمت نمایشی', 0) or 0)),
                            price_per_unit=int(product_data.get('قیمت به ازای', 0) or 0),
                            has_quantity=product_data.get('دارای تیراژ', 'خیر') == 'بله',
                            is_active=product_data.get('فعال', 'بله') == 'بله',
                        )
                        product.save()
                        
                        # ذخیره نگاشت ID
                        if product_id_col:
                            temp_id = row[product_id_col]
                            temp_to_real_ids[f"product_{temp_id}"] = product.id
                        
                        imported_count += 1
                        
                    except Exception as e:
                        failed_count += 1
                        errors.append(f"سطر {row_num} (محصول): {str(e)}")
                        logger.error(f"Error importing product at row {row_num}: {str(e)}")
                
                if failed_count > 0:
                    raise Exception("خطا در ایجاد محصولات" + str(failed_count))
                
                # ===== مرحله ۲: ایجاد فیلدها =====
                if fields_ws:
                    field_headers = [cell.value for cell in fields_ws[1]]
                    product_id_col = field_headers.index('Product ID') if 'Product ID' in field_headers else None
                    field_id_col = field_headers.index('Field ID') if 'Field ID' in field_headers else None
                    
                    temp_field_ids = {}  # نگاشت temp_id به ID واقعی فیلد
                    
                    for row_num, row in enumerate(fields_ws.iter_rows(min_row=2, values_only=True), 2):
                        if not any(row):
                            continue
                        
                        try:
                            field_data = dict(zip(field_headers, row))
                            product_temp_id = field_data.get('Product ID')
                            
                            # پیدا کردن محصول
                            product_id = temp_to_real_ids.get(f"product_{product_temp_id}")
                            if not product_id:
                                errors.append(f"سطر {row_num} (فیلد): محصول با ID {product_temp_id} یافت نشد.")
                                continue
                            
                            # ایجاد یا به‌روزرسانی FieldDictionary
                            field_dict, _ = FieldDictionary.objects.get_or_create(
                                title=field_data.get('Field Title', '').strip(),
                                defaults={
                                    'field_type': field_data.get('Field Type', 'dropdown'),
                                    'description': field_data.get('Description', '') or '',
                                }
                            )
                            
                            # ایجاد ProductField
                            product_field = ProductField.objects.create(
                                product_id=product_id,
                                field_dict=field_dict,
                                numeric_value=Decimal(str(field_data.get('Value', 0) or 0)),
                                is_required=field_data.get('Required', 'خیر') == 'بله',
                                is_active=True,
                                order=int(field_data.get('Order', 0) or 0)
                            )
                            
                            # ذخیره نگاشت
                            if field_id_col:
                                temp_field_id = row[field_id_col]
                                temp_field_ids[temp_field_id] = product_field.id
                            
                        except Exception as e:
                            errors.append(f"سطر {row_num} (فیلد): {str(e)}")
                            logger.error(f"Error importing field at row {row_num}: {str(e)}")
                
                # ===== مرحله ۳: ایجاد فرمول‌ها =====
                if formulas_ws:
                    formula_headers = [cell.value for cell in formulas_ws[1]]
                    product_id_col = formula_headers.index('Product ID') if 'Product ID' in formula_headers else None
                    
                    for row_num, row in enumerate(formulas_ws.iter_rows(min_row=2, values_only=True), 2):
                        if not any(row):
                            continue
                        
                        try:
                            formula_data = dict(zip(formula_headers, row))
                            product_temp_id = formula_data.get('Product ID')
                            
                            # پیدا کردن محصول
                            product_id = temp_to_real_ids.get(f"product_{product_temp_id}")
                            if not product_id:
                                errors.append(f"سطر {row_num} (فرمول): محصول با ID {product_temp_id} یافت نشد.")
                                continue
                            
                            # جایگزینی ID فیلدها در عبارت‌ها
                            calc_expr = formula_data.get('Calculation Expression', '') or ''
                            cond_expr = formula_data.get('Condition', '') or ''
                            
                            # جایگزینی field_{temp_id} با field_{real_id}
                            for temp_id, real_id in temp_field_ids.items():
                                calc_expr = calc_expr.replace(f"field_{temp_id}", f"field_{real_id}")
                                cond_expr = cond_expr.replace(f"field_{temp_id}", f"field_{real_id}")
                            
                            # ایجاد فرمول
                            ProductFormula.objects.create(
                                product_id=product_id,
                                title=formula_data.get('Title', '').strip(),
                                condition_expression=cond_expr if cond_expr else None,
                                calculation_expression=calc_expr
                            )
                            
                        except Exception as e:
                            errors.append(f"سطر {row_num} (فرمول): {str(e)}")
                            logger.error(f"Error importing formula at row {row_num}: {str(e)}")
                
                # ===== مرحله ۴: دانلود و آپلود عکس‌ها =====
                if images_ws:
                    image_headers = [cell.value for cell in images_ws[1]]
                    product_id_col = image_headers.index('Product ID') if 'Product ID' in image_headers else None
                    image_url_col = image_headers.index('Image URL') if 'Image URL' in image_headers else None
                    
                    for row_num, row in enumerate(images_ws.iter_rows(min_row=2, values_only=True), 2):
                        if not any(row):
                            continue
                        
                        try:
                            image_data = dict(zip(image_headers, row))
                            product_temp_id = image_data.get('Product ID')
                            image_url = image_data.get('Image URL')
                            
                            if not image_url:
                                continue
                            
                            # پیدا کردن محصول
                            product_id = temp_to_real_ids.get(f"product_{product_temp_id}")
                            if not product_id:
                                errors.append(f"سطر {row_num} (عکس): محصول با ID {product_temp_id} یافت نشد.")
                                continue
                            
                            # دانلود عکس از URL
                            import requests
                            response = requests.get(image_url, timeout=30)
                            if response.status_code == 200:
                                ext = image_url.split('.')[-1].split('?')[0]
                                file_name = f"products/{product_id}/{row_num}_{int(timezone.now().timestamp())}.{ext}"
                                
                                # ذخیره در default_storage
                                default_storage.save(file_name, ContentFile(response.content))
                                
                                # ایجاد رکورد ProductImage
                                ProductImage.objects.create(
                                    product_id=product_id,
                                    user=user,
                                    image=file_name,
                                    order=row_num
                                )
                            
                        except Exception as e:
                            errors.append(f"سطر {row_num} (عکس): {str(e)}")
                            logger.error(f"Error importing image at row {row_num}: {str(e)}")
            
            logger.info(f"Import completed: {imported_count} imported, {failed_count} failed")
            
            return {
                'success': True,
                'message': f'ایمپورت تکمیم شد: {imported_count} موفق، {failed_count} ناموفق.',
                'imported_count': imported_count,
                'failed_count': failed_count,
                'errors': errors
            }
            
        except Exception as e:
            logger.error(f"Error importing products: {str(e)}", exc_info=True)
            return {
                'success': False,
                'message': f'خطا در ایمپورت محصولات: {str(e)}',
                'imported_count': 0,
                'failed_count': 0,
                'errors': [str(e)]
            }
    
    # ==========================================
    # EXPORT WITH IMAGES & ATTACHMENTS
    # ==========================================
    
    def export_products_to_excel_enhanced(self, product_ids: Optional[List[int]] = None) -> Dict[str, Any]:
        """
        استخراج محصولات به Excel شامل عکس‌ها و پیوست‌ها
        
        Args:
            product_ids: لیست ID محصولات (اگر None یا خالی باشد، همه محصولات استخراج می‌شوند)
        
        Returns:
            Dict containing file info
        """
        try:
            # دریافت محصولات
            if product_ids and len(product_ids) > 0:
                products = Product.objects.filter(
                    id__in=product_ids
                ).prefetch_related(
                    'category_relations__category',
                    'fields__field_dict',
                    'fields__choices__choice_dict',
                    'formulas',
                    'product_image',
                    'product_attachment'
                )
            else:
                products = Product.objects.all().prefetch_related(
                    'category_relations__category',
                    'fields__field_dict',
                    'fields__choices__choice_dict',
                    'formulas',
                    'product_image',
                    'product_attachment'
                )
            
            products = products.order_by('-created_at')
            product_count = products.count()
            
            if product_count == 0:
                return {
                    'success': False,
                    'message': 'هیچ محصولی برای استخراج یافت نشد.',
                    'file_path': None,
                    'product_count': 0
                }
            
            # ایجاد Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # استایل‌ها
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ===== Sheet 1: محصولات =====
            headers = [
                "ID", "نام محصول", "کد محصول", "توضیحات", "قیمت", 
                "قیمت نمایشی", "قیمت به ازای", "دارای تیراژ", "فعال",
                "دسته‌بندی‌ها", "تاریخ ایجاد"
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            for row_num, product in enumerate(products, 2):
                categories = self._get_product_categories_text(product)
                
                row_data = [
                    product.id,
                    product.name,
                    product.code or "",
                    product.description or "",
                    float(product.price) if product.price else 0,
                    float(product.show_price) if product.show_price else 0,
                    product.price_per_unit or 0,
                    "بله" if product.has_quantity else "خیر",
                    "بله" if product.is_active else "خیر",
                    categories,
                    product.created_at.strftime('%Y-%m-%d %H:%M') if product.created_at else ""
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = cell_alignment
                    cell.border = border
            
            column_widths = [8, 30, 20, 50, 15, 15, 15, 12, 10, 40, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 2: فیلدهای محصولات =====
            if products.exists():
                ws_fields = wb.create_sheet("Product Fields")
                field_headers = [
                    "Product ID", "Product Name", "Field ID", "Field Title", 
                    "Field Type", "Value", "Required", "Order"
                ]
                
                for col_num, header in enumerate(field_headers, 1):
                    cell = ws_fields.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                row_num = 2
                for product in products:
                    for field in product.fields.all():
                        field_data = [
                            product.id,
                            product.name,
                            field.id,
                            field.field_dict.title,
                            field.field_dict.field_type,
                            float(field.numeric_value) if field.numeric_value is not None else 0,
                            "بله" if field.is_required else "خیر",
                            field.order
                        ]
                        
                        for col_num, value in enumerate(field_data, 1):
                            cell = ws_fields.cell(row=row_num, column=col_num)
                            cell.value = value
                            cell.alignment = cell_alignment
                            cell.border = border
                        row_num += 1
                
                field_column_widths = [12, 30, 10, 25, 15, 15, 10, 8]
                for i, width in enumerate(field_column_widths, 1):
                    ws_fields.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 3: فرمول‌ها =====
            if products.exists():
                ws_formulas = wb.create_sheet("Formulas")
                formula_headers = [
                    "Product ID", "Product Name", "Formula ID", "Title",
                    "Condition", "Calculation Expression"
                ]
                
                for col_num, header in enumerate(formula_headers, 1):
                    cell = ws_formulas.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                row_num = 2
                for product in products:
                    for formula in product.formulas.all():
                        formula_data = [
                            product.id,
                            product.name,
                            formula.id,
                            formula.title,
                            formula.condition_expression or "",
                            formula.calculation_expression
                        ]
                        
                        for col_num, value in enumerate(formula_data, 1):
                            cell = ws_formulas.cell(row=row_num, column=col_num)
                            cell.value = value
                            cell.alignment = cell_alignment
                            cell.border = border
                        row_num += 1
                
                formula_column_widths = [12, 30, 10, 25, 30, 50]
                for i, width in enumerate(formula_column_widths, 1):
                    ws_formulas.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 4: عکس‌ها =====
            if products.exists():
                ws_images = wb.create_sheet("Images")
                image_headers = [
                    "Product ID", "Product Name", "Image ID", "Image URL",
                    "Order", "Created At"
                ]
                
                for col_num, header in enumerate(image_headers, 1):
                    cell = ws_images.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                row_num = 2
                for product in products:
                    for image in product.product_image.all():
                        image_data = [
                            product.id,
                            product.name,
                            image.id,
                            image.image.url if image.image else "",
                            image.order,
                            image.created_at.strftime('%Y-%m-%d %H:%M') if image.created_at else ""
                        ]
                        
                        for col_num, value in enumerate(image_data, 1):
                            cell = ws_images.cell(row=row_num, column=col_num)
                            cell.value = value
                            cell.alignment = cell_alignment
                            cell.border = border
                        row_num += 1
                
                image_column_widths = [12, 30, 10, 80, 8, 20]
                for i, width in enumerate(image_column_widths, 1):
                    ws_images.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 5: فایل‌های پیوست =====
            if products.exists():
                ws_attachments = wb.create_sheet("Attachments")
                attachment_headers = [
                    "Product ID", "Product Name", "Attachment ID", "File Name",
                    "File URL", "Created At"
                ]
                
                for col_num, header in enumerate(attachment_headers, 1):
                    cell = ws_attachments.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                row_num = 2
                for product in products:
                    for attachment in product.product_attachment.all():
                        attachment_data = [
                            product.id,
                            product.name,
                            attachment.id,
                            attachment.name or "",
                            attachment.file.url if attachment.file else "",
                            attachment.created_at.strftime('%Y-%m-%d %H:%M') if attachment.created_at else ""
                        ]
                        
                        for col_num, value in enumerate(attachment_data, 1):
                            cell = ws_attachments.cell(row=row_num, column=col_num)
                            cell.value = value
                            cell.alignment = cell_alignment
                            cell.border = border
                        row_num += 1
                
                attachment_column_widths = [12, 30, 10, 30, 80, 20]
                for i, width in enumerate(attachment_column_widths, 1):
                    ws_attachments.column_dimensions[get_column_letter(i)].width = width
            
            # ذخیره فایل
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"products_export_{timestamp}.xlsx"
            file_path = f"exports/products/{file_name}"
            
            exports_dir = os.path.join(default_storage.location, 'exports', 'products')
            os.makedirs(exports_dir, exist_ok=True)
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
            default_storage.save(file_path, buffer)
            
            logger.info(f"Successfully exported {product_count} products to {file_path}")
            
            return {
                'success': True,
                'message': f'{product_count} محصول با موفقیت استخراج شد.',
                'file_path': file_path,
                'file_name': file_name,
                'product_count': product_count
            }
            
        except Exception as e:
            logger.error(f"Error exporting products: {str(e)}", exc_info=True)
            return {
                'success': False,
                'message': f'خطا در استخراج محصولات: {str(e)}',
                'file_path': None,
                'product_count': 0
            }
    
    # ==========================================
    # GET EXPORT TEMPLATE (Enhanced)
    # ==========================================
    
    def get_export_template(self) -> Dict[str, Any]:
        """
        ایجاد فایل Excel نمونه برای ایمپورت (همراه با فیلدها، فرمول‌ها و عکس‌ها)
        
        Returns:
            Dict containing:
                - file_path: مسیر فایل نمونه
                - file_name: نام فایل
        """
        try:
            wb = Workbook()
            
            # ===== Sheet 1: محصولات =====
            ws_products = wb.active
            ws_products.title = "Products"
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            product_headers = [
                "ID", "نام محصول *", "توضیحات", "قیمت", "قیمت نمایشی",
                "قیمت به ازای (تیراژ مبنای)", "دارای تیراژ", "فعال"
            ]
            
            for col_num, header in enumerate(product_headers, 1):
                cell = ws_products.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            sample_products = [
                [1, "کارت ویزیت لمینت", "چاپ افست با کیفیت بالا", 20000, 100000, 1000, "بله", "بله"],
                [2, "ست اداری", "شامل پوشه و سربرگ", 50000, 250000, 500, "بله", "بله"],
            ]
            
            for row_num, row_data in enumerate(sample_products, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_products.cell(row=row_num, column=col_num, value=value)
            
            product_column_widths = [8, 30, 50, 15, 15, 20, 12, 10]
            for i, width in enumerate(product_column_widths, 1):
                ws_products.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 2: فیلدهای محصولات =====
            ws_fields = wb.create_sheet("Product Fields")
            field_headers = [
                "Product ID", "Field ID", "Field Title", "Field Type",
                "Value", "Required", "Order"
            ]
            
            for col_num, header in enumerate(field_headers, 1):
                cell = ws_fields.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            sample_fields = [
                [1, 1, "نوع کاغذ", "dropdown", 0, "بله", 1],
                [1, 2, "تعداد رنگ", "single_select", 0, "بله", 2],
                [2, 3, "جنس", "dropdown", 0, "خیر", 1],
            ]
            
            for row_num, row_data in enumerate(sample_fields, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_fields.cell(row=row_num, column=col_num, value=value)
            
            field_column_widths = [12, 10, 25, 15, 15, 10, 8]
            for i, width in enumerate(field_column_widths, 1):
                ws_fields.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 3: فرمول‌ها =====
            ws_formulas = wb.create_sheet("Formulas")
            formula_headers = [
                "Product ID", "Formula ID", "Title", "Condition",
                "Calculation Expression"
            ]
            
            for col_num, header in enumerate(formula_headers, 1):
                cell = ws_formulas.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            sample_formulas = [
                [1, 1, "فرمول استاندارد", "", "(field_1 * field_2) + 5000"],
                [2, 2, "فرمول ویژه", "field_3 == 50", "(field_3 * 1000) + 10000"],
            ]
            
            for row_num, row_data in enumerate(sample_formulas, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_formulas.cell(row=row_num, column=col_num, value=value)
            
            formula_column_widths = [12, 10, 25, 30, 50]
            for i, width in enumerate(formula_column_widths, 1):
                ws_formulas.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 4: عکس‌ها =====
            ws_images = wb.create_sheet("Images")
            image_headers = [
                "Product ID", "Image ID", "Image URL", "Order"
            ]
            
            for col_num, header in enumerate(image_headers, 1):
                cell = ws_images.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            sample_images = [
                [1, 1, "https://example.com/images/product1.jpg", 1],
                [1, 2, "https://example.com/images/product1_2.jpg", 2],
            ]
            
            for row_num, row_data in enumerate(sample_images, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_images.cell(row=row_num, column=col_num, value=value)
            
            image_column_widths = [12, 10, 80, 8]
            for i, width in enumerate(image_column_widths, 1):
                ws_images.column_dimensions[get_column_letter(i)].width = width
            
            # ===== Sheet 5: راهنما =====
            ws_help = wb.create_sheet("راهنما")
            help_text = [
                ["راهنمای کامل ایمپورت:", ""],
                ["", ""],
                ["1. Products (محصولات):", ""],
                ["ID", "شناسه محصول (در ایمپورت جدید، این فیلد خالی بمانید)"],
                ["نام محصول *", "نام محصول (اجباری)"],
                ["توضیحات", "توضیحات محصول"],
                ["قیمت", "قیمت پایه (عدد صحیح)"],
                ["قیمت نمایشی", "قیمت نمایش در سایت"],
                ["قیمت به ازای", "تعداد مبنا (مثلاً ۱۰۰۰)"],
                ["دارای تیراژ", "بله/خیر"],
                ["فعال", "بله/خیر"],
                ["", ""],
                ["2. Product Fields (فیلدهای داینامیک):", ""],
                ["Product ID", "شناسه محصول از Sheet محصولات"],
                ["Field ID", "شناسه فیلد (در ایمپورت جدید، خالی بمانید)"],
                ["Field Title", "نام فیلد (مثال: نوع کاغذ)"],
                ["Field Type", "نوع فیلد: text, textarea, number, single_select, multi_select, dropdown"],
                ["Value", "مقدار عددی پایه"],
                ["Required", "بله/خیر - آیا اجباری است؟"],
                ["Order", "ترتیب نمایش"],
                ["", ""],
                ["3. Formulas (فرمول‌های قیمت‌گذاری):", ""],
                ["Product ID", "شناسه محصول"],
                ["Formula ID", "شناسه فرمول (در ایمپورت جدید، خالی بمانید)"],
                ["Title", "عنوان فرمول"],
                ["Condition", "شرط اجرای فرمول (مثال: field_1 > 1000)"],
                ["Calculation Expression", "عبارت محاسباتی (مثال: field_1 * 1000)"],
                ["", ""],
                ["4. Images (عکس‌ها):", ""],
                ["Product ID", "شناسه محصول"],
                ["Image ID", "شناسه عکس (در ایمپورت جدید، خالی بمانید)"],
                ["Image URL", "آدرس کامل عکس (https://...)"],
                ["Order", "ترتیب نمایش"],
                ["", ""],
                ["نکات مهم:", ""],
                ["- IDهای خالی به صورت خودکار تولید می‌شوند", ""],
                ["- در فرمول‌ها از ID واقعی فیلدها استفاده کنید", ""],
                ["- عکس‌ها از URL دانلود و آپلود می‌شوند", ""],
                ["- فیلدهای اجباری: نام محصول", ""],
            ]
            
            for row_num, row_data in enumerate(help_text, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws_help.cell(row=row_num, column=col_num)
                    cell.value = value
                    if col_num == 1 and ":" in str(value):
                        cell.font = Font(bold=True)
            
            ws_help.column_dimensions['A'].width = 25
            ws_help.column_dimensions['B'].width = 70
            
            # ذخیره فایل
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"products_import_template_{timestamp}.xlsx"
            file_path = f"templates/products/{file_name}"
            
            templates_dir = os.path.join(default_storage.location, 'templates', 'products')
            os.makedirs(templates_dir, exist_ok=True)
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
            default_storage.save(file_path, buffer)
            
            return {
                'success': True,
                'file_path': file_path,
                'file_name': file_name
            }
            
        except Exception as e:
            logger.error(f"Error creating template: {str(e)}", exc_info=True)
            return {
                'success': False,
                'message': f'خطا در ایجاد فایل نمونه: {str(e)}',
                'file_path': None,
                'file_name': None
            }
